import csv
import openpyxl
import xlsxwriter
from B2B_clients import B2B_clients_dict
from B2C_clients import B2C_clients_dict
from chat_clients import chat_clients_dict

def read_NBA_WNBA(file_name):

    wb = openpyxl.load_workbook(file_name) 
    #get names of all tabs in the spreadsheet
    tab_names_list = wb.sheetnames
    #return the worksheet objects for each of those names
    worksheets = []
    for tab_name in tab_names_list:
        worksheets.append(wb.get_sheet_by_name(tab_name))

    #return worksheet objects which then feeds into filter_NBA_WNBA
    return {'NBA': worksheets[0], 'WNBA': worksheets[1]}


def read_daily_stats(file_name):

    wb = openpyxl.load_workbook(file_name) 
    #get names of all tabs in the spreadsheet
    tab_names_list = wb.sheetnames
    #return the worksheet objects for each of those names
    worksheets = []
    for tab_name in tab_names_list:
        worksheets.append(wb.get_sheet_by_name(tab_name))

    #return worksheet objects which then feeds into filter_daily_stats
    return {'Current Backlog': worksheets[0], 'Daily Support Contacts': worksheets[1], 
            'Previous Day P1P2 Ticket...': worksheets[2]} 


def filter_NBA_WNBA(NBA_WNBA_worksheets_dict):
    
    NBA_emails = 0
    NBA_chats = 0
    WNBA_emails = 0

    #iterate through the items in the rows to add up values of emails and seperate them from chats
    for row in NBA_WNBA_worksheets_dict['NBA'].iter_rows(values_only=True):
        for item in row:
            if 'Email' == item:
                NBA_emails += int(float(row[2]))
            elif 'Web' == item:
                NBA_emails += int(float(row[2]))
            elif 'WhatsApp' == item:
                NBA_emails += int(float(row[2]))
            elif 'Messaging' == item:
                NBA_chats += int(float(row[2]))

    #same as above but only need the total figure ('SUM')
    for row in NBA_WNBA_worksheets_dict['WNBA'].iter_rows(values_only=True):
        for item in row:
            if 'SUM' == item:
                WNBA_emails += int(float(row[2]))
                break
                # break as there are two 'SUM' items in the row (don't want to add twice)

    #feeds into filter_daily_stats in order to combine the NBA and WNBA stats
    return {'NBA - email': NBA_emails, 'NBA - chat': NBA_chats, 'WNBA - email': WNBA_emails}


def filter_daily_stats(daily_stats_worksheets_dict, NBA_WNBA_stats_dict):


    #NEW:
    print('\n New \n')
    #seperate B2B from B2C tickets
    B2B = []
    B2C = []
    for row in daily_stats_worksheets_dict['Daily Support Contacts'].iter_rows(values_only=True):
        print(row)
        if row[0] == 'B2B':
            B2B.append(row)
        elif row[0] == 'B2C':
            B2C.append(row)


    print('\n B2B \n')
    print(B2B)
    #B2B New:

    #make new copy of B2B_clients_dict
    B2B_email = B2B_clients_dict.copy()
    #for every client name (row[2], find it in B2B_email and add the original value (0) with the ticket count value (row[-1] - 'SUM'))
    for row in B2B: 
        try: 
            B2B_email[row[2]]+=int(row[-1])
        except KeyError:
            print(f'KeyError thrown for {row[2]} (May not exist in B2B_clients_dict)')
            continue


    print('\n B2B_email \n')
    print(B2B_email)
    print(len(B2B_email))

    print('\n B2C \n')
    print(B2C)

    #B2C New:

    #Similar process as above but also combine the NBA and WNBA stats AND seperate emails from chats

    #Emails:
    B2C_email = B2C_clients_dict.copy()
    for row in B2C: 
        try:          
            if isinstance(row[3], float):
                B2C_email[row[2]] += int(row[3])
            if isinstance(row[5], float):
                B2C_email[row[2]] += int(row[5])
            if isinstance(row[7], float):
                B2C_email[row[2]] += int(row[7])
        except KeyError:
            print(f'KeyError thrown for {row[2]} (May not exist in B2C_clients_dict)')
            continue

        #in order to add 'Any Channel', 'Email' and 'Web' values together.
        #need to check if the value is a float (e.g. 1.0) first, as if the value is nothing,
        #it's actually an empty string (e.g. '') and will throw an error if you try to operate.

    #Combine the NBA and WNBA email stats
    B2C_email['NBA']=NBA_WNBA_stats_dict['NBA - email']
    B2C_email['WNBA']=NBA_WNBA_stats_dict['WNBA - email']

    print('\n B2C_email \n')
    print(B2C_email)
    print(len(B2C_email))

    #B2C Chat:
    B2C_chat = chat_clients_dict.copy()
    for row in B2C: 
        try:          
            if isinstance(row[4], float):
                B2C_chat[row[2]] += int(row[4])
            if isinstance(row[6], float):
                B2C_chat[row[2]] += int(row[6])
        except KeyError:
            print(f'KeyError thrown for {row[2]} (May not exist in chat_clients_dict)')
            continue

        #in order to add 'Chat' and 'Messaging' values together.
        #need to check if the value is a float (e.g. 1.0) first, as if the value is nothing,
        #it's actually an empty string (e.g. '') and will throw an error if you try to operate.

    print('\n B2C_chat \n')
    print(B2C_chat)
    print(len(B2C_chat))


    #Backlog:
    print('\nBacklog: \n')
    #seperate B2B from B2C tickets
    B2B = []
    B2C = []
    for row in daily_stats_worksheets_dict['Current Backlog'].iter_rows(values_only=True):
        print(row)
        if row[0] == 'B2B':
            B2B.append(row)
        #'type' cell for Sky, NESN and WWE B2B backlog can appear as '\xa0' (non-breaking space)
        elif row[0] == '\xa0':
            B2B.append(row)
        elif row[0] == 'B2C':
            B2C.append(row)

    print('\n B2B \n')
    print(B2B)
    #B2B Backlog:

    #make new copy of B2B_clients_dict
    B2B_backlog = B2B_clients_dict.copy()
    #for every client name (row[1], find it in B2B_backlog and add the original value (0) with the ticket count value (row[-1]))
    for row in B2B: 
        try: 
            B2B_backlog[row[1]]+=int(row[-1])
        except KeyError:
            print(f'KeyError thrown for {row[1]} (May not exist in B2B_clients_dict)')
            continue


    print('\n B2B_backlog \n')
    print(B2B_backlog)
    print(len(B2B_backlog))


    #B2C Backlog:
    print('\n B2C \n')
    print(B2C)

    #Same as above but for B2C

    B2C_backlog = B2C_clients_dict.copy()
    for row in B2C: 
        try: 
            B2C_backlog[row[1]]+=int(row[-1])
        except KeyError:
            print(f'KeyError thrown for {row[1]} (May not exist in B2C_clients_dict)')
            continue


    print('\n B2C_backlog \n')
    print(B2C_backlog)
    print(len(B2C_backlog))
    


    #return a dictionary with the key representing each tab name for the new xlsx file and the value a dict
    #with the field name and row values for those tabs.
    #The xlsx file replicates the layout of the 'Daily Team Stats' spreadsheet (The tabs with quantative data only).
    return {'B2B Email': B2B_email, 'B2B Backlog': B2B_backlog, 'B2C Email': B2C_email, 'B2C Backlog': B2C_backlog, 'B2C Chat': B2C_chat}



def create_new_xlsx_file(rows_dict):
    #set row and column index values to 0 (0,0 = A1 cell)
    row = 0
    col = 0
    #create the new xlsx file
    workbook = xlsxwriter.Workbook('Daily_Team_Stats.xlsx')

    #Add a worksheet/tab for each of the keys in the rows_dict. The key names are the tab names.
    B2B_email = workbook.add_worksheet('B2B Email')
    #add client names along the top row and the values in the row below
    for client, value in rows_dict['B2B Email'].items():
        B2B_email.write(row, col, client)
        B2B_email.write(row + 1, col, value)
        col += 1
    #reset column index to 0 for the next tab
    col = 0
    print('B2B_email complete')

    B2B_backlog = workbook.add_worksheet('B2B Backlog')

    for client, value in rows_dict['B2B Backlog'].items():
        B2B_backlog.write(row, col, client)
        B2B_backlog.write(row + 1, col, value)
        col += 1

    col = 0
    print('B2B_backlog complete')

    B2C_email = workbook.add_worksheet('B2C Email')

    for client, value in rows_dict['B2C Email'].items():
        B2C_email.write(row, col, client)
        B2C_email.write(row + 1, col, value)
        col += 1

    col = 0
    print('B2C_email complete')

    B2C_backlog = workbook.add_worksheet('B2C Backlog')

    for client, value in rows_dict['B2C Backlog'].items():
        B2C_backlog.write(row, col, client)
        B2C_backlog.write(row + 1, col, value)
        col += 1

    col = 0
    print('B2C_backlog complete')

    B2C_chat = workbook.add_worksheet('B2C Chat')

    for client, value in rows_dict['B2C Chat'].items():
        B2C_chat.write(row, col, client)
        B2C_chat.write(row + 1, col, value)
        col += 1

    col = 0
    print('B2C_chat complete')

    workbook.close()